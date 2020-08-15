# -*- coding: utf-8 -*-
"""
Created on Fri Aug 14 23:18:16 2020

@author: kagimub
"""



#loan status graph
rows = dataframe_to_rows(dash2)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx+7, column=c_idx+2, value=value)
         
#draw charts
from openpyxl.chart import BarChart, Reference, Series
"""chart = BarChart()
values = Reference(ws, min_col=3, min_row=10, max_col=4, max_row=20)
chart.add_data(values)
ws.add_chart(chart, "E15")

chart = BarChart()
xvalues = Reference(ws, min_col=3, min_row=10, max_row=20)
for i in range(4, 5):
    values = Reference(ws, min_col=i, min_row=10, max_row=20)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)


chart.title = "Default layout"
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'
chart.legend.position = 'r'

ws.add_chart(chart, "E10")"""



row_count = ws3.max_row
column_count = ws3.max_column

cols=['B','C','D','E','F','G','H','I','J','K']

for column in cols:
    for row in range(2, row_count):
        ws3["{0}{1}".format(column,row)].value = float(ws3["{0}{1}".format(column,row)].value)
        ws3["{0}{1}".format(column,row)].number_format = '#,##0'



